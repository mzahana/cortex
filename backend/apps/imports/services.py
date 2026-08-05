"""Parsing/validation/commit engine for the CSV/Excel bulk importer (T6.1).

**ASSUMPTION flagged (Q8, `docs/risks.md`): no representative spreadsheet
sample was provided.** This module is designed against a generic, reasonable
asset-spreadsheet schema instead — the following columns, matched
case-insensitively and whitespace-trimmed against a spreadsheet's header row:

    name        (required)  — `Asset.name`.
    category    (required)  — matched BY NAME against the tenant's `Category`
                               tree (`apps.catalog.models.Category`, a
                               self-referential parent/child tree). Matching
                               is a case-insensitive exact match on `name`
                               ANYWHERE in the tree — a name that is not
                               unique tenant-wide (two categories with the
                               same name under different parents) is reported
                               as an "ambiguous category name" row error
                               rather than silently picking one; a lab's
                               category names are expected to be
                               tenant-wide-unique in practice even though the
                               model itself only enforces uniqueness per
                               (parent, name).
    location    (optional)  — same by-name matching against the `Location`
                               tree; blank = no location.
    project     (optional)  — by-name match against `Project` (tenant-wide
                               unique by `Project.name`, so never ambiguous);
                               blank = general pool.
    status      (optional)  — matched case-insensitively against
                               `Asset.Status` choices; defaults to
                               `Asset.Status.AVAILABLE` when blank/omitted.
    condition   (optional)  — free text, stored as-is on `Asset.condition`.
    tags        (optional)  — comma-separated tag names; matched-or-created
                               by name (`Tag.objects.get_or_create`), same as
                               `AssetSerializer._sync_tags` does for a normal
                               `POST /assets`.

Every OTHER column header is treated as a **custom-field column**: for each
row, once `category` has resolved, the header is matched case-insensitively
against that category's `CustomFieldDef.key` or `.label`; the cell value is
then validated/coerced by the SAME `apps.assets.services.
validate_custom_field_values` used by ordinary asset create/edit (reused
verbatim here, not reimplemented — a `text|int|float|bool|date|enum|json`
type-checked value, `required` enforcement, `enum_options` membership all
apply identically). A column that matches no custom field for a given row's
category is simply ignored for that row (it may still apply to a different
row whose category defines a field with that key/label) — no ambiguity error
is raised for this case, unlike category/location, because two DIFFERENT
categories legitimately define fields under the same spreadsheet column
(e.g. a "power_watts" column used by both "Compute" and "Edge" rows).

This exact schema is also what `apps.imports.exports.stream_assets_csv`
exports — see that module for the round-trip: an export always emits
`name, category, location, status, condition, project, tags` plus one column
per DISTINCT custom-field KEY used by the exported assets' categories, so
re-importing an unmodified export reproduces equivalent assets.

**Streaming, bounded-RAM parsing (R2, CLAUDE.md "slow work runs in
Celery... chunked to bound worker RAM").** `iter_source_rows` never
materializes an entire spreadsheet in memory: CSV is read row-by-row via the
stdlib `csv` module over the storage file's own stream; `.xlsx` is opened
with `openpyxl.load_workbook(..., read_only=True)`, which streams worksheet
rows from disk rather than loading the whole workbook into memory (the
`pandas.read_excel`-style "load it all" approach this deliberately avoids).
`resolve_import_rows` is itself a generator over that stream, so a dry-run
report is built with peak memory proportional to ONE row at a time during
parsing, converging to O(row count) only when the caller (`build_report`)
collects the per-row report list for storage — reasonable and bounded for a
lab inventory spreadsheet (M1 already proved 10k+ assets is comfortably
fast/well-indexed; a spreadsheet describing that many rows of everyday lab
gear, at a few hundred bytes of resolved state each, is a few MB, not a
worker-RAM risk).

**Commit is all-or-nothing under a single transaction (documented decision,
task's own preferred default for MVP).** `commit_import_rows` fully resolves
every row FIRST (still via the same streaming/bounded parse above) and only
opens `transaction.atomic()` if EVERY row is valid; if any row is invalid it
creates nothing and returns the same per-row report so the caller sees
exactly what's still wrong. This is simpler and safer than a partial-commit
("valid rows only") for a spreadsheet-sized dataset — a half-imported
inventory sheet with silently-skipped rows is a worse failure mode for a lab
than "fix the 3 flagged rows and re-run the whole file", and it does not
conflict with the bounded-RAM requirement above: chunking there is about the
PARSE step never loading the whole file at once, not about the number of
transactions the DB write uses (a lab inventory import is at most low
thousands of rows — a single transaction's rowset, not a memory problem).
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from typing import Any, Iterable, Iterator

import openpyxl
from django.core.files.storage import default_storage
from django.db import transaction
from rest_framework import serializers

from apps.assets.models import Asset, AssetFieldValue, TagLink
from apps.assets.services import validate_custom_field_values
from apps.catalog.models import Category, CustomFieldDef, Location, Tag
from apps.projects.models import Project

# --- Column-mapping targets (module docstring is the schema authority) ------

TARGET_NAME = "name"
TARGET_CATEGORY = "category"
TARGET_LOCATION = "location"
TARGET_PROJECT = "project"
TARGET_STATUS = "status"
TARGET_CONDITION = "condition"
TARGET_TAGS = "tags"
TARGET_URL = "url"
TARGET_CUSTOM = "custom"
TARGET_IGNORE = "ignore"

CORE_TARGETS: frozenset[str] = frozenset(
    {
        TARGET_NAME,
        TARGET_CATEGORY,
        TARGET_LOCATION,
        TARGET_PROJECT,
        TARGET_STATUS,
        TARGET_CONDITION,
        TARGET_TAGS,
        # `Asset.url` is a CORE target, not a custom field: it must round-trip
        # through `apps.imports.exports.CORE_EXPORT_COLUMNS` (which includes
        # it). If it were left unmapped, `default_column_mapping` would fall
        # through to `TARGET_CUSTOM` and a re-imported export would try to
        # materialize the built-in column as a per-category custom field.
        TARGET_URL,
    }
)
ALL_TARGETS: frozenset[str] = CORE_TARGETS | {TARGET_CUSTOM, TARGET_IGNORE}

SUPPORTED_EXTENSIONS: frozenset[str] = frozenset({"csv", "xlsx"})


def file_extension(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def default_column_mapping(headers: list[str]) -> dict[str, str]:
    """Auto-match every header case-insensitively/trimmed against
    `CORE_TARGETS`; anything unmatched defaults to `"custom"` (per-row/
    per-category custom-field resolution, see module docstring) rather than
    `"ignore"` — every column is used unless the caller explicitly opts a
    header out via an override mapping.
    """
    mapping: dict[str, str] = {}
    for header in headers:
        key = header.strip().lower()
        mapping[header] = key if key in CORE_TARGETS else TARGET_CUSTOM
    return mapping


def resolve_column_mapping(headers: list[str], override: dict[str, str] | None) -> dict[str, str]:
    """The auto-detected default (`default_column_mapping`), with any
    caller-supplied `override` entries replacing the default for the SAME
    header (an override for a header that isn't actually present in the
    file is silently ignored — it can't apply to anything).
    """
    resolved = default_column_mapping(headers)
    if override:
        for header, target in override.items():
            if header in resolved and target in ALL_TARGETS:
                resolved[header] = target
    return resolved


# --- Streaming source parsing (R2: bounded RAM) ------------------------------


def iter_source_rows(fileobj: Any, filename: str) -> Iterator[tuple[list[str], dict[str, Any]]]:
    """Yields `(headers, row_dict)` for every DATA row (the header row itself
    is consumed first and not yielded as data). `row_dict` is
    `{header: cell_value}`; missing trailing cells map to `None`.

    Streams from `fileobj` row-by-row for both formats — see module
    docstring's "bounded RAM" note.
    """
    ext = file_extension(filename)
    if ext == "csv":
        text_stream = io.TextIOWrapper(fileobj, encoding="utf-8-sig", newline="")
        reader = csv.reader(text_stream)
        try:
            headers = next(reader)
        except StopIteration:
            return
        headers = [h.strip() for h in headers]
        for raw_row in reader:
            if not any(cell.strip() for cell in raw_row if cell):
                continue  # skip a fully-blank line
            padded = raw_row + [None] * (len(headers) - len(raw_row))
            yield headers, dict(zip(headers, padded, strict=True))
        return

    if ext == "xlsx":
        workbook = openpyxl.load_workbook(fileobj, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            if worksheet is None:  # an empty workbook with no sheets at all
                return
            rows_iter = worksheet.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                return
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            for raw_row in rows_iter:
                if all(cell is None for cell in raw_row):
                    continue
                padded = list(raw_row) + [None] * (len(headers) - len(raw_row))
                yield headers, dict(zip(headers, padded, strict=True))
        finally:
            workbook.close()
        return

    raise ValueError(f"Unsupported file extension '.{ext}'. Only .csv and .xlsx are supported.")


def _normalize_cell(value: Any) -> Any:
    """Cell -> a plain Python value ready for `_coerce_value`-style
    validation: blank strings and `None` both become `None` ("not supplied");
    `datetime.date`/`datetime.datetime` (native openpyxl cell types)
    normalize to an ISO date string; everything else passes through as-is
    (CSV cells are already plain `str`; openpyxl already gives native
    `int`/`float`/`bool`/`str` for the rest).
    """
    import datetime

    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped != "" else None
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()[:10]
    return value


# --- Row resolution -----------------------------------------------------------


@dataclass
class ResolvedRow:
    row_number: int  # 1-based, counts header as row 1 (so row 2 = first data row)
    name: str | None = None
    category: Category | None = None
    location: Location | None = None
    project: Project | None = None
    tag_names: list[str] = field(default_factory=list)
    status: str = Asset.Status.AVAILABLE
    condition: str = ""
    url: str = ""
    custom_field_pairs: list[tuple[CustomFieldDef, Any]] = field(default_factory=list)
    # `Any`, not `str`: most entries are a plain message string, but
    # `custom_field_values` nests `validate_custom_field_values`'s own
    # `{key: message}` dict verbatim (see below).
    errors: dict[str, Any] = field(default_factory=dict)

    @property
    def is_valid(self) -> bool:
        return not self.errors

    def to_report_dict(self) -> dict[str, Any]:
        return {
            "row_number": self.row_number,
            "values": {
                "name": self.name,
                "category": self.category.name if self.category else None,
                "location": self.location.name if self.location else None,
                "project": self.project.name if self.project else None,
                "tags": self.tag_names,
                "status": self.status,
                "condition": self.condition,
                "url": self.url,
                "custom_field_values": {fd.key: value for fd, value in self.custom_field_pairs},
            },
            "errors": self.errors,
        }


class _NameIndex:
    """Case-insensitive-name -> object index over a tenant's tree/list, with
    ambiguity detection (module docstring: two rows sharing a `name` at
    different places in the `Category`/`Location` tree is reported as an
    error, never silently resolved to "the first match").
    """

    def __init__(self, objects: Iterable[Any]):
        self._by_name: dict[str, list[Any]] = {}
        for obj in objects:
            self._by_name.setdefault(obj.name.strip().lower(), []).append(obj)

    def resolve(self, raw_name: str) -> tuple[Any | None, str | None]:
        """Returns `(object_or_none, error_or_none)`."""
        matches = self._by_name.get(raw_name.strip().lower(), [])
        if not matches:
            return None, f"'{raw_name}' does not match any known name."
        if len(matches) > 1:
            return None, f"'{raw_name}' is ambiguous ({len(matches)} matches)."
        return matches[0], None


def resolve_import_rows(
    *,
    tenant,
    source_stream: Any,
    filename: str,
    mapping_override: dict[str, str] | None,
    resolved_mapping_out: dict[str, str] | None = None,
) -> Iterator[ResolvedRow]:
    """Streams `ResolvedRow`s for every data row in `source_stream`
    (`filename` picks the CSV/xlsx parser). `tenant` is used to build the
    Category/Location/Project name indexes ONCE up front (small, tenant-wide
    admin-config-sized queries — not per-row) — caller MUST already be
    inside `apps.tenancy.context.tenant_context(tenant.id)` so the plain
    `Category.objects`/`Location.objects`/`Project.objects`/`Tag.objects`
    tenant-scoped managers resolve correctly (R4).

    `mapping_override` is resolved against the file's ACTUAL header row
    (`resolve_column_mapping`, module docstring) the moment it's read — a
    single streaming pass, never a separate "peek the header row first"
    pass (which would need a second `default_storage.open()`/`seek(0)`
    dance that's fragile across storage backends). If `resolved_mapping_out`
    (a plain `dict`) is supplied, it is mutated in place with the resolved
    `{header: target}` mapping as soon as it's known, so a caller that fully
    drains this generator (both `build_report` and `commit_import_rows` do)
    can read the ACTUAL confirmed mapping back out afterward without a
    second parse.
    """
    categories = _NameIndex(Category.objects.all())
    locations = _NameIndex(Location.objects.all())
    projects = _NameIndex(Project.objects.all())
    status_by_lower = {choice.lower(): choice for choice in Asset.Status.values}

    header_to_target: dict[str, str] | None = None
    row_number = 1  # header row is row 1

    for headers, raw_row in iter_source_rows(source_stream, filename):
        if header_to_target is None:
            header_to_target = resolve_column_mapping(headers, mapping_override)
            if resolved_mapping_out is not None:
                resolved_mapping_out.update(header_to_target)
        row_number += 1

        resolved = ResolvedRow(row_number=row_number)
        custom_raw: dict[str, Any] = {}

        for header, cell in raw_row.items():
            target = header_to_target.get(header, TARGET_CUSTOM)
            value = _normalize_cell(cell)

            if target == TARGET_IGNORE:
                continue
            if target == TARGET_NAME:
                resolved.name = value
            elif target == TARGET_CATEGORY:
                if value is not None:
                    obj, err = categories.resolve(str(value))
                    if err:
                        resolved.errors[TARGET_CATEGORY] = err
                    else:
                        resolved.category = obj
            elif target == TARGET_LOCATION:
                if value is not None:
                    obj, err = locations.resolve(str(value))
                    if err:
                        resolved.errors[TARGET_LOCATION] = err
                    else:
                        resolved.location = obj
            elif target == TARGET_PROJECT:
                if value is not None:
                    obj, err = projects.resolve(str(value))
                    if err:
                        resolved.errors[TARGET_PROJECT] = err
                    else:
                        resolved.project = obj
            elif target == TARGET_STATUS:
                if value is not None:
                    resolved_status = status_by_lower.get(str(value).strip().lower())
                    if resolved_status is None:
                        resolved.errors[TARGET_STATUS] = (
                            f"'{value}' is not a valid status "
                            f"({', '.join(Asset.Status.values)})."
                        )
                    else:
                        resolved.status = resolved_status
            elif target == TARGET_CONDITION:
                resolved.condition = str(value) if value is not None else ""
            elif target == TARGET_URL:
                raw_url = str(value).strip() if value is not None else ""
                if raw_url and not raw_url.lower().startswith(("http://", "https://")):
                    # Same http/https-only rule the API enforces
                    # (`apps.assets.serializers.AssetSerializer.validate_url`)
                    # -- a spreadsheet is just another write path into the
                    # same field, and the value ends up in an `<a href>`.
                    resolved.errors[TARGET_URL] = "URL must start with http:// or https://."
                else:
                    resolved.url = raw_url
            elif target == TARGET_TAGS:
                if value is not None:
                    resolved.tag_names = [
                        part.strip() for part in str(value).split(",") if part.strip()
                    ]
            else:  # TARGET_CUSTOM
                custom_raw[header.strip().lower()] = value

        if resolved.name is None:
            resolved.errors[TARGET_NAME] = "This field is required."
        if resolved.category is None and TARGET_CATEGORY not in resolved.errors:
            # Covers BOTH "column present but blank" and "no header maps to
            # category at all" — either way it's required.
            resolved.errors[TARGET_CATEGORY] = "This field is required."

        # Custom fields: only resolvable once `category` is known — matched
        # per-row against THAT category's field defs, by key or label
        # (case-insensitive), per the module docstring.
        if resolved.category is not None and custom_raw:
            field_defs = list(resolved.category.field_defs.all())
            by_key = {fd.key.lower(): fd for fd in field_defs}
            by_label = {fd.label.strip().lower(): fd for fd in field_defs}

            values_by_key: dict[str, Any] = {}
            for header_lower, raw_value in custom_raw.items():
                field_def = by_key.get(header_lower) or by_label.get(header_lower)
                if field_def is None:
                    continue  # column doesn't apply to this row's category
                values_by_key[field_def.key] = raw_value

            # Also require any field the category defines but whose column
            # wasn't present at all in this file, so `required=True` custom
            # fields are still enforced (mirrors ordinary asset creation).
            try:
                pairs = validate_custom_field_values(resolved.category, values_by_key)
            except serializers.ValidationError as exc:
                detail = exc.detail
                if isinstance(detail, dict):
                    errors = detail.get("custom_field_values", detail)
                else:
                    errors = detail
                resolved.errors["custom_field_values"] = errors
            else:
                resolved.custom_field_pairs = pairs

        yield resolved


# --- Report building (dry-run + commit's own re-validation) -----------------


def build_report(
    *, tenant, source_stream: Any, filename: str, mapping_override: dict[str, str] | None
) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    valid_count = 0
    invalid_count = 0
    resolved_mapping: dict[str, str] = {}
    for resolved in resolve_import_rows(
        tenant=tenant,
        source_stream=source_stream,
        filename=filename,
        mapping_override=mapping_override,
        resolved_mapping_out=resolved_mapping,
    ):
        rows.append(resolved.to_report_dict())
        if resolved.is_valid:
            valid_count += 1
        else:
            invalid_count += 1
    return {
        "resolved_mapping": resolved_mapping,
        "rows": rows,
        "total_rows": len(rows),
        "valid_count": valid_count,
        "invalid_count": invalid_count,
    }


# --- Commit -------------------------------------------------------------------


def commit_import_rows(
    *, tenant, source_stream: Any, filename: str, mapping_override: dict[str, str] | None
) -> tuple[list[int], dict[str, Any]]:
    """All-or-nothing (module docstring): fully resolves every row first; if
    any row is invalid, creates nothing and returns `([], report)`. If every
    row is valid, creates one `Asset` (+ custom-field values + tags) per row
    inside a single `transaction.atomic()` block and returns
    `(created_asset_ids, report)`.

    Caller must already be inside `tenant_context(tenant.id)` (same
    requirement as `resolve_import_rows`) — `Asset.objects`/`Tag.objects`
    etc. below are the tenant-scoped managers.
    """
    resolved_mapping: dict[str, str] = {}
    resolved_rows = list(
        resolve_import_rows(
            tenant=tenant,
            source_stream=source_stream,
            filename=filename,
            mapping_override=mapping_override,
            resolved_mapping_out=resolved_mapping,
        )
    )
    rows_report = [r.to_report_dict() for r in resolved_rows]
    valid_count = sum(1 for r in resolved_rows if r.is_valid)
    invalid_count = len(resolved_rows) - valid_count
    report = {
        "resolved_mapping": resolved_mapping,
        "rows": rows_report,
        "total_rows": len(resolved_rows),
        "valid_count": valid_count,
        "invalid_count": invalid_count,
    }

    if invalid_count or not resolved_rows:
        return [], report

    created_ids: list[int] = []
    with transaction.atomic():
        tag_cache: dict[str, Tag] = {}
        for row in resolved_rows:
            # Every row here is already `is_valid` (the `invalid_count`
            # early-return above), and `category` is a required field
            # (`resolve_import_rows` always adds an error for a missing/
            # unresolved category) -- so `row.category` is guaranteed
            # non-`None` at this point; asserted for the type checker.
            assert row.category is not None
            asset = Asset.objects.create(
                tenant=tenant,
                category=row.category,
                name=row.name,
                project=row.project,
                location=row.location,
                status=row.status,
                condition=row.condition,
                url=row.url,
                is_consumable=row.category.default_is_consumable,
            )
            if row.custom_field_pairs:
                AssetFieldValue.objects.bulk_create(
                    AssetFieldValue(tenant=tenant, asset=asset, field_def=field_def, value=value)
                    for field_def, value in row.custom_field_pairs
                )
            for tag_name in row.tag_names:
                key = tag_name.strip()
                if not key:
                    continue
                tag = tag_cache.get(key.lower())
                if tag is None:
                    tag, _ = Tag.objects.get_or_create(tenant=tenant, name=key)
                    tag_cache[key.lower()] = tag
                TagLink.objects.create(tenant=tenant, asset=asset, tag=tag)
            created_ids.append(asset.id)

    return created_ids, report


# --- Source file storage ------------------------------------------------------


def import_source_storage_key(tenant_id: int, import_job_id: int, filename: str) -> str:
    safe_name = filename.replace("/", "_").replace("\\", "_")
    return f"imports/{tenant_id}/{import_job_id}/{safe_name}"


def save_import_source_file(
    *, tenant_id: int, import_job_id: int, uploaded_file: Any
) -> tuple[str, str, str]:
    """Writes `uploaded_file`'s bytes to the storage backend, same
    bytes-never-touch-the-DB pattern as `apps.assets.services.
    save_attachment_file`. Returns `(storage_key, filename, content_type)`.
    """
    key = import_source_storage_key(tenant_id, import_job_id, uploaded_file.name)
    storage_key = default_storage.save(key, uploaded_file)
    content_type = getattr(uploaded_file, "content_type", "") or ""
    return storage_key, uploaded_file.name, content_type
